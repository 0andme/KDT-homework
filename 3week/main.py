'''
미니프로젝트 (3).ipynb 파일을 처음 보시는 분을 위해!
​
(1) 프로젝트 명 : 네이버 뉴스 자동 수집 후 메일 발송기
​
(2) 프로젝트 설명(200자 이내) : 특정 키워드에 해당하는 뉴스 기사를 네이버에서 찾아 기사 제목, 링크를 엑셀 파일로 만들어 미리 만들어둔 엑셀 파일에 있는 메일링 대상자에게 메일로 전달합니다.
​
(3) 프로젝트 과제 상세 :
​
사용자가 원하는 키워드 입력받기
네이버 뉴스를 수집해 주는 모듈을 이용해서 해당 키워드 뉴스 수집 후 엑셀 파일에 제목, 링크, 요약문 기록하기
수집 데이터 엑셀 파일을 메일링 대상자 엑셀 파일을 읽어 대상자들에게 메일 보내기
(4) 점검 및 합격 기준표 :
​
해당 파트를 수강하면서 사용한 문법과 기능을 활용할 수 있어야 합니다.
전체적인 업무의 흐름을 이해하고 작성하여야 합니다.
각 단계마다 주어진 과제 코드를 완성하여야 합니다.
​
​
아래 코드를 실행해서 NaverNewsCrawler 모듈을 임포트하세요.
​
# 크롤러 코드를 위한 모듈 설치
!pip install requests
!pip install beautifulsoup4
'''

from NaverNewsCrawler import NaverNewsCrawler

####사용자로 부터 기사 수집을 원하는 키워드를 input을 이용해 입력받기
userKeyword=input("뉴스 키워드를 입력하세요 ex) 패스트 캠퍼스 : ")
crawler = NaverNewsCrawler(userKeyword)

#### 수집한 데이터를 저장할 엑셀 파일명을 input을 이용해 입력받기
userfileName=input("수집한 데이터를 저장할 엑셀 파일명을 입력하세요 ex) 파일명.xlsx : ")
crawler.get_news(userfileName)

#### 이메일 발송 기능에 필요한 모듈을 임포트
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re

#### gmail 발송 기능에 필요한 계정 정보
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
#### 개인 정보 이기에 github push시 제외
SMTP_USER = ''
SMTP_PASSWORD = ''

#### 메일 발송에 필요한 send_mail 함수
def send_mail(name, addr, subject, contents, attachment):
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative')
    if attachment:
        msg = MIMEMultipart('mixed')

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="' + filename + '"')
        msg.attach(file_data)

    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()



#### 프로젝트 폴더에 있는 email_list.xlsx 파일에 이메일 받을 사람들의 정보를 입력
# email_list.xlsx에 두 사람의 data를 저장
# 1, '유영미', '21_smilebom@naver.com'
# 2, '홍길동', 'yu00mi97@gmail.com'

#### 엑셀 파일의 정보를 읽어올 수 있는 모듈을 import

from openpyxl import load_workbook

#### email_list.xlsx 파일을 읽어와 해당 사람들에게 수집한 뉴스 정보 엑셀 파일
#### end_mail 함수를 이용해 전송
# 엑셀 파일 읽기
wb = load_workbook('email_list.xlsx', read_only=True)
data = wb.active

startRow = 3          # startRow : 메일링 대상자들의 이름 data가 시작되는 행 저장 변수
maxRow = data.max_row # maxRow : 엑셀의 data가 끝나는 행 저장 변수
# 메일 sub 제목
sub_title= userKeyword +'에 대한 뉴스 정보 메일입니다' 

# 엑셀 파일을 startRow 부터 maxRow까지 한 행씩 읽어 for문 실행
for row in data[startRow:maxRow]:
      send_mail(
        row[1].value,
        row[2].value,
        sub_title,
        userKeyword+'에 대한 뉴스 정보를 '+ row[1].value+'님께 엑셀 파일로 보내드립니다',
        userfileName )

        